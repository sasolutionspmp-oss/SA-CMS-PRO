from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
try:  # pragma: no cover - python-docx optional
    from docx import Document
except Exception:  # pragma: no cover
    Document = None
try:  # pragma: no cover - pillow optional
    from PIL import Image
except Exception:  # pragma: no cover
    Image = None


def parse_pdf(path: Path) -> str:
    try:
        from pypdf import PdfReader
        reader = PdfReader(str(path))
        return "".join(page.extract_text() or "" for page in reader.pages)
    except Exception:  # pragma: no cover
        return ""


def parse_docx(path: Path) -> str:
    if not Document:
        return ""
    doc = Document(str(path))
    return "\n".join(p.text for p in doc.paragraphs)


def parse_xlsx(path: Path) -> str:
    wb = load_workbook(str(path), read_only=True)
    texts = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    texts.append(str(cell))
    return "\n".join(texts)


def parse_txt(path: Path) -> str:
    return path.read_text()


def parse_image(path: Path) -> str:
    if not Image:  # pragma: no cover
        return ""
    try:
        import pytesseract  # lazy import
    except Exception:  # pragma: no cover
        return ""
    try:
        img = Image.open(path)
        return pytesseract.image_to_string(img)
    except Exception:
        return ""


EXT_PARSERS = {
    '.pdf': parse_pdf,
    '.docx': parse_docx,
    '.xlsx': parse_xlsx,
    '.txt': parse_txt,
    '.png': parse_image,
    '.jpg': parse_image,
    '.jpeg': parse_image,
    '.tiff': parse_image,
}


def parse_file(path: Path) -> str:
    func = EXT_PARSERS.get(path.suffix.lower())
    if not func:
        return ""
    return func(path)
