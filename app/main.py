import io
import zipfile
from pathlib import Path
from typing import List
from uuid import uuid4

from fastapi import FastAPI, UploadFile
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse

from .db import init_db, get_db
from . import parsing, chunking, fts
from .jobs import create_job, log, update_status, get_job

app = FastAPI()


@app.on_event("startup")
def startup():
    init_db()


@app.post("/api/ingest/zip")
async def ingest_zip(file: UploadFile):
    job_id = create_job("ingest")
    update_status(job_id, "running")
    dest = Path("uploads") / f"{uuid4()}.zip"
    dest.parent.mkdir(parents=True, exist_ok=True)
    with dest.open("wb") as f:
        while chunk := await file.read(1024 * 1024):
            f.write(chunk)
    process_zip(dest, job_id)
    return {"job_id": job_id}


def process_zip(zip_path: Path, job_id: int):
    try:
        log(job_id, "extract")
        extract_dir = Path("extracts") / zip_path.stem
        extract_dir.mkdir(parents=True, exist_ok=True)
        with zipfile.ZipFile(zip_path) as z:
            z.extractall(extract_dir)
        for path in extract_dir.rglob("*"):
            if path.is_file():
                log(job_id, f"parse {path.name}")
                text = parsing.parse_file(path)
                conn = get_db()
                cur = conn.cursor()
                cur.execute(
                    "INSERT INTO files(path, kind, size, source_zip) VALUES (?,?,?,?)",
                    (str(path), path.suffix.lower(), path.stat().st_size, str(zip_path)),
                )
                file_id = cur.lastrowid
                if text:
                    chunk_dir = Path("chunks") / f"{file_id}"
                    chunks = chunking.chunk_text(text, chunk_dir)
                    for idx, cpath in enumerate(chunks):
                        cur.execute(
                            "INSERT INTO chunks(file_id, idx, text_path, page_hint, size) VALUES (?,?,?,?,?)",
                            (file_id, idx, str(cpath), None, cpath.stat().st_size),
                        )
                    conn.commit()
                    fts.index_chunks(file_id, chunks)
                else:
                    conn.commit()
        update_status(job_id, "done")
    except Exception as e:  # pragma: no cover
        log(job_id, f"error {e}")
        update_status(job_id, "error")


@app.get("/api/jobs/{job_id}")
def job_status(job_id: int):
    job = get_job(job_id)
    if not job:
        return JSONResponse(status_code=404, content={"detail": "not found"})
    return job


@app.get("/api/ingest/sample")
def sample_zip():
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w') as z:
        # PDF
        from fpdf import FPDF
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        pdf.cell(200, 10, txt="hello pdf", ln=True)
        pdf_bytes = pdf.output(dest='S').encode('latin1')
        z.writestr('sample.pdf', pdf_bytes)
        # XLSX
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'hello xlsx'
        xlsx_bytes = io.BytesIO()
        wb.save(xlsx_bytes)
        z.writestr('sample.xlsx', xlsx_bytes.getvalue())
        # PNG
        from PIL import Image, ImageDraw
        img = Image.new('RGB', (200, 60), color=(255,255,255))
        d = ImageDraw.Draw(img)
        d.text((10,10), 'hello img', fill=(0,0,0))
        img_bytes = io.BytesIO()
        img.save(img_bytes, format='PNG')
        z.writestr('sample.png', img_bytes.getvalue())
    buf.seek(0)
    headers = {'Content-Disposition': 'attachment; filename="sample.zip"'}
    return StreamingResponse(buf, media_type='application/zip', headers=headers)
